from fastapi import FastAPI, HTTPException, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.utils import get_column_letter
import io

app = FastAPI(title="Forecast Management API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_PATH = Path(__file__).parent.parent.parent / "forecast_dummy.xlsx"

MONTH_COLUMNS = {
    "Oct": {"on_rev": "Total Oct ON Rev", "on_hc": "Total Oct ON HC", "off_rev": "Total Oct Off Rev", "off_hc": "Total Oct Off HC"},
    "Nov": {"on_rev": "Total Nov ON Rev", "on_hc": "Total Nov ON HC", "off_rev": "Total Nov Off Rev", "off_hc": "Total Nov Off HC"},
    "Dec": {"on_rev": "Total Dec ON Rev", "on_hc": "Total Dec ON HC", "off_rev": "Total Dec Off Rev", "off_hc": "Total Dec Off HC"},
    "Jan": {"on_rev": "Total Jan ON Rev", "on_hc": "Total Jan ON HC", "off_rev": "Total Jan Off Rev", "off_hc": "Total Jan Off HC"},
    "Feb": {"on_rev": "Total Feb ON Rev", "on_hc": "Total Feb ON HC", "off_rev": "Total Feb Off Rev", "off_hc": "Total Feb Off HC"},
    "Mar": {"on_rev": "Total Mar ON Rev", "on_hc": "Total Mar ON HC", "off_rev": "Total Mar Off Rev", "off_hc": "Total Mar Off HC"},
    "Apr": {"on_rev": "Total Apr ON Rev", "on_hc": "Total Apr ON HC", "off_rev": "Total Apr Off Rev", "off_hc": "Total Apr Off HC"},
    "May": {"on_rev": "Total May ON Rev", "on_hc": "Total May ON HC", "off_rev": "Total May Off Rev", "off_hc": "Total May Off HC"},
    "Jun": {"on_rev": "Total Jun ON Rev", "on_hc": "Total Jun ON HC", "off_rev": "Total Jun Off Rev", "off_hc": "Total Jun Off HC"},
    "Jul": {"on_rev": "Total Jul ON Rev", "on_hc": "Total Jul ON HC", "off_rev": "Total Jul Off Rev", "off_hc": "Total Jul Off HC"},
    "Aug": {"on_rev": "Total Aug ON Rev", "on_hc": "Total Aug ON HC", "off_rev": "Total Aug Off Rev", "off_hc": "Total Aug Off HC"},
    "Sep": {"on_rev": "Total Sep ON Rev", "on_hc": "Total Sep ON HC", "off_rev": "Total Sep Off Rev", "off_hc": "Total Sep Off HC"},
}

QUARTER_COLUMNS = {
    "OND": {"on_rev": "Total OND ON Rev", "on_hc": "Total OND ON HC", "off_rev": "Total OND Off Rev", "off_hc": "Total OND Off HC"},
    "JFM": {"on_rev": "Total JFM ON Rev", "on_hc": "Total JFM ON HC", "off_rev": "Total JFM Off Rev", "off_hc": "Total JFM Off HC"},
    "AMJ": {"on_rev": "Total AMJ ON Rev", "on_hc": "Total AMJ ON HC", "off_rev": "Total AMJ Off Rev", "off_hc": "Total AMJ Off HC"},
    "JAS": {"on_rev": "Total JAS ON Rev", "on_hc": "Total JAS ON HC", "off_rev": "Total JAS Off Rev", "off_hc": "Total JAS Off HC"},
}

class MonthlyData(BaseModel):
    on_rev: Optional[float] = 0
    on_hc: Optional[float] = 0
    off_rev: Optional[float] = 0
    off_hc: Optional[float] = 0

class QuarterlyData(BaseModel):
    on_rev: Optional[float] = 0
    on_hc: Optional[float] = 0
    off_rev: Optional[float] = 0
    off_hc: Optional[float] = 0

class Project(BaseModel):
    id: int
    division: str
    client: str
    project: str
    bu: str
    coach: str
    project_type: str
    monthly: Dict[str, MonthlyData]
    quarterly: Dict[str, QuarterlyData]

class ProjectUpdate(BaseModel):
    monthly: Dict[str, MonthlyData]
    quarterly: Dict[str, QuarterlyData]

def safe_float(value):
    if pd.isna(value) or value == '' or value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

@app.get("/api/projects", response_model=List[Project])
async def get_projects():
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Sheet1")
        
        projects = []
        for idx, row in df.iterrows():
            monthly_data = {}
            for month, cols in MONTH_COLUMNS.items():
                monthly_data[month] = MonthlyData(
                    on_rev=safe_float(row.get(cols["on_rev"])),
                    on_hc=safe_float(row.get(cols["on_hc"])),
                    off_rev=safe_float(row.get(cols["off_rev"])),
                    off_hc=safe_float(row.get(cols["off_hc"]))
                )
            
            quarterly_data = {}
            for quarter, cols in QUARTER_COLUMNS.items():
                quarterly_data[quarter] = QuarterlyData(
                    on_rev=safe_float(row.get(cols["on_rev"])),
                    on_hc=safe_float(row.get(cols["on_hc"])),
                    off_rev=safe_float(row.get(cols["off_rev"])),
                    off_hc=safe_float(row.get(cols["off_hc"]))
                )
            
            project = Project(
                id=idx,
                division=str(row.get("Division", "")),
                client=str(row.get("Client name", "")),
                project=str(row.get("Project name", "")),
                bu=str(row.get("BU", "")),
                coach=str(row.get("Project coach", "")),
                project_type=str(row.get("Project Type", "")),
                monthly=monthly_data,
                quarterly=quarterly_data
            )
            projects.append(project)
        
        return projects
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading Excel file: {str(e)}")

def calculate_quarterly_totals(monthly_data: Dict[str, MonthlyData]) -> Dict[str, QuarterlyData]:
    quarter_months = {
        'OND': ['Oct', 'Nov', 'Dec'],
        'JFM': ['Jan', 'Feb', 'Mar'],
        'AMJ': ['Apr', 'May', 'Jun'],
        'JAS': ['Jul', 'Aug', 'Sep']
    }
    
    quarterly_totals = {}
    for quarter, months in quarter_months.items():
        on_rev = sum(monthly_data.get(m, MonthlyData()).on_rev for m in months)
        on_hc = sum(monthly_data.get(m, MonthlyData()).on_hc for m in months)
        off_rev = sum(monthly_data.get(m, MonthlyData()).off_rev for m in months)
        off_hc = sum(monthly_data.get(m, MonthlyData()).off_hc for m in months)
        
        quarterly_totals[quarter] = QuarterlyData(
            on_rev=on_rev,
            on_hc=on_hc,
            off_rev=off_rev,
            off_hc=off_hc
        )
    
    return quarterly_totals

@app.put("/api/projects/{project_id}")
async def update_project(project_id: int, update: ProjectUpdate):
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Sheet1")
        if project_id >= len(df):
            raise HTTPException(status_code=404, detail="Project not found")
        
        row = df.iloc[project_id]
        
        all_monthly_data = {}
        for month in ['Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep']:
            if month in update.monthly:
                all_monthly_data[month] = update.monthly[month]
            else:
                cols = MONTH_COLUMNS.get(month, {})
                on_rev_col = cols.get('on_rev', '')
                on_hc_col = cols.get('on_hc', '')
                off_rev_col = cols.get('off_rev', '')
                off_hc_col = cols.get('off_hc', '')
                
                all_monthly_data[month] = MonthlyData(
                    on_rev=safe_float(row.get(on_rev_col, 0) if on_rev_col else 0),
                    on_hc=safe_float(row.get(on_hc_col, 0) if on_hc_col else 0),
                    off_rev=safe_float(row.get(off_rev_col, 0) if off_rev_col else 0),
                    off_hc=safe_float(row.get(off_hc_col, 0) if off_hc_col else 0)
                )
        
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Sheet1"]
        row_num = project_id + 2
        
        col_map = {}
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_map[col_name] = get_column_letter(col_idx)
        
        for month, data in all_monthly_data.items():
            if month in MONTH_COLUMNS:
                cols = MONTH_COLUMNS[month]
                ws[f"{col_map[cols['on_rev']]}{row_num}"] = data.on_rev
                ws[f"{col_map[cols['on_hc']]}{row_num}"] = data.on_hc
                ws[f"{col_map[cols['off_rev']]}{row_num}"] = data.off_rev
                ws[f"{col_map[cols['off_hc']]}{row_num}"] = data.off_hc
        
        calculated_quarterly = calculate_quarterly_totals(all_monthly_data)
        
        for quarter, data in calculated_quarterly.items():
            if quarter in QUARTER_COLUMNS:
                cols = QUARTER_COLUMNS[quarter]
                ws[f"{col_map[cols['on_rev']]}{row_num}"] = data.on_rev
                ws[f"{col_map[cols['on_hc']]}{row_num}"] = data.on_hc
                ws[f"{col_map[cols['off_rev']]}{row_num}"] = data.off_rev
                ws[f"{col_map[cols['off_hc']]}{row_num}"] = data.off_hc
        
        wb.save(EXCEL_PATH)
        wb.close()
        
        return {"status": "success", "message": f"Project {project_id} updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating Excel file: {str(e)}")

@app.get("/api/dashboard")
async def get_dashboard_data(coach: Optional[str] = None, project_name: Optional[str] = None):
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Sheet1")
        
        coaches = ['All'] + sorted(df['Project coach'].dropna().unique().tolist())
        
        filtered_df = df.copy()
        if coach and coach != 'All':
            filtered_df = filtered_df[filtered_df['Project coach'] == coach]
        
        project_names = ['All'] + sorted(filtered_df['Project name'].dropna().unique().tolist())
        
        if project_name and project_name != 'All':
            filtered_df = filtered_df[filtered_df['Project name'] == project_name]
        
        dashboard_data = {
            "quarters": ["OND", "JFM", "AMJ", "JAS"],
            "on_hc": [],
            "on_rev": [],
            "off_hc": [],
            "off_rev": [],
            "total_hc": [],
            "total_rev": [],
            "coaches": coaches,
            "project_names": project_names
        }
        
        df = filtered_df
        
        for quarter in ['OND', 'JFM', 'AMJ', 'JAS']:
            if quarter in QUARTER_COLUMNS:
                cols = QUARTER_COLUMNS[quarter]
                
                on_rev_col = cols.get('on_rev', '')
                on_hc_col = cols.get('on_hc', '')
                off_rev_col = cols.get('off_rev', '')
                off_hc_col = cols.get('off_hc', '')
                
                on_rev = df[on_rev_col].sum() if on_rev_col in df.columns else 0
                on_hc = df[on_hc_col].sum() if on_hc_col in df.columns else 0
                off_rev = df[off_rev_col].sum() if off_rev_col in df.columns else 0
                off_hc = df[off_hc_col].sum() if off_hc_col in df.columns else 0
                
                dashboard_data['on_hc'].append(round(float(on_hc), 2))
                dashboard_data['on_rev'].append(round(float(on_rev), 2))
                dashboard_data['off_hc'].append(round(float(off_hc), 2))
                dashboard_data['off_rev'].append(round(float(off_rev), 2))
                dashboard_data['total_hc'].append(round(float(on_hc + off_hc), 2))
                dashboard_data['total_rev'].append(round(float(on_rev + off_rev), 2))
        
        return dashboard_data
    except Exception as e:
        import traceback
        error_detail = f"Error generating dashboard data: {str(e)}\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=error_detail)

@app.post("/api/compare")
async def compare_files(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        uploaded_df = pd.read_excel(io.BytesIO(contents), sheet_name="Sheet1")
        current_df = pd.read_excel(EXCEL_PATH, sheet_name="Sheet1")
        
        project_comparisons = []
        
        if 'Project name' not in current_df.columns:
            raise HTTPException(status_code=400, detail="Project name column not found")
        
        for idx, current_row in current_df.iterrows():
            project_name = current_row.get('Project name', '')
            uploaded_row = uploaded_df[uploaded_df['Project name'] == project_name]
            
            if uploaded_row.empty:
                continue
            
            uploaded_row = uploaded_row.iloc[0]
            
            project_changes = {
                'project_name': project_name,
                'client_name': current_row.get('Client name', ''),
                'project_coach': current_row.get('Project coach', ''),
                'quarters': []
            }
            
            has_changes = False
            
            for quarter in ['OND', 'JFM', 'AMJ', 'JAS']:
                if quarter not in QUARTER_COLUMNS:
                    continue
                
                cols = QUARTER_COLUMNS[quarter]
                
                quarter_data = {
                    'quarter': quarter,
                    'changes': []
                }
                
                for metric_type, col_name in [
                    ('On HC', cols.get('on_hc', '')),
                    ('On Rev', cols.get('on_rev', '')),
                    ('Off HC', cols.get('off_hc', '')),
                    ('Off Rev', cols.get('off_rev', ''))
                ]:
                    if col_name and col_name in current_df.columns and col_name in uploaded_df.columns:
                        current_val = float(current_row.get(col_name, 0)) if pd.notna(current_row.get(col_name)) else 0
                        uploaded_val = float(uploaded_row.get(col_name, 0)) if pd.notna(uploaded_row.get(col_name)) else 0
                        
                        if abs(current_val - uploaded_val) > 0.01:
                            quarter_data['changes'].append({
                                'metric': metric_type,
                                'current': round(current_val, 2),
                                'uploaded': round(uploaded_val, 2),
                                'difference': round(uploaded_val - current_val, 2)
                            })
                            has_changes = True
                
                if quarter_data['changes']:
                    project_changes['quarters'].append(quarter_data)
            
            if has_changes:
                project_comparisons.append(project_changes)
        
        return {
            'status': 'success',
            'total_projects_with_changes': len(project_comparisons),
            'total_projects_current': len(current_df),
            'total_projects_uploaded': len(uploaded_df),
            'current_file_name': EXCEL_PATH.name,
            'uploaded_file_name': file.filename,
            'projects': project_comparisons
        }
    except Exception as e:
        import traceback
        error_detail = f"Error comparing files: {str(e)}\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=error_detail)

@app.get("/health")
async def health_check():
    return {"status": "healthy", "excel_path": str(EXCEL_PATH), "exists": EXCEL_PATH.exists()}
